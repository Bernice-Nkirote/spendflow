import uuid
from io import BytesIO
from uuid import UUID

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook

from app.models.supplier import Supplier
from app.repositories.supplier_repository import SupplierRepository
from app.schemas.supplier_schema import SupplierImportResult
from app.services.audit_log_service import AuditLogService
from app.services.permission_service import PermissionService

HEADER_ALIASES = {
    "name": {
        "name",
        "supplier",
        "supplier_name",
        "supplier company",
        "company name",
    },
    "email": {
        "email",
        "supplier_email",
        "email_address",
        "email address",
    },
    "phone": {
        "phone",
        "telephone",
        "mobile",
        "mobile_number",
        "phone_number",
        "phone number",
    },
    "address": {
        "address",
        "location",
        "physical_address",
        "physical address",
    },
    "contact_person": {
        "contact_person",
        "contact person",
        "contact",
        "representative",
        "supplier_contact",
    },
}


class SupplierExcelImportService:

    def __init__(
        self,
        repo: SupplierRepository,
        audit_log_service: AuditLogService,
        permission_service: PermissionService,
    ):
        self.repo = repo
        self.audit_log_service = audit_log_service
        self.permission_service = permission_service

    async def import_suppliers_from_excel(
        self,
        file: UploadFile,
        company_id: UUID,
        actor_user_id: UUID,
        actor_role_id: UUID,
    ) -> SupplierImportResult:
        if not self.permission_service.role_has_permission(
            role_id=actor_role_id,
            permission_name="suppliers.import",
            company_id=company_id,
        ):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have permission to import suppliers.",
            )

        self._validate_file(file)

        contents = await file.read()
        sheet = self._load_excel_sheet(contents)

        header_indexes = self._get_header_indexes(sheet)

        if sheet.max_row < 2:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "The uploaded Excel file has headers but no supplier rows. "
                    "Add supplier data starting from row 2 and upload again."
                ),
            )

        created_suppliers: list[Supplier] = []
        errors: list[dict] = []

        seen_names: set[str] = set()
        seen_emails: set[str] = set()
        seen_phones: set[str] = set()

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            raw_data = self._extract_row_data(row, header_indexes)

            row_error = self._validate_row(
                row_number=row_number,
                raw_data=raw_data,
                company_id=company_id,
                seen_names=seen_names,
                seen_emails=seen_emails,
                seen_phones=seen_phones,
            )

            if row_error:
                errors.append(row_error)
                continue

            supplier = self._build_supplier(
                raw_data=raw_data,
                company_id=company_id,
            )

            created_supplier = self.repo.create(supplier)
            created_suppliers.append(created_supplier)

            seen_names.add(supplier.name.lower())

            if supplier.email:
                seen_emails.add(supplier.email)

            if supplier.phone:
                seen_phones.add(supplier.phone)

        self.audit_log_service.log_action(
            company_id=company_id,
            entity_type="SUPPLIER",
            entity_id=created_suppliers[0].id if created_suppliers else company_id,
            action="SUPPLIERS_IMPORTED",
            actor_user_id=actor_user_id,
            description=(
                f"Supplier Excel import completed. "
                f"{len(created_suppliers)} supplier(s) created and {len(errors)} row(s) failed."
            ),
            details_json={
                "entity_reference": "Supplier Excel Import",
                "created_count": len(created_suppliers),
                "failed_count": len(errors),
                "file_name": file.filename,
            },
        )

        self.repo.db.commit()

        for supplier in created_suppliers:
            self.repo.db.refresh(supplier)

        if len(created_suppliers) == 0 and len(errors) > 0:
            errors.insert(
                0,
                {
                    "row": 0,
                    "message": (
                        "No new suppliers were imported. It looks like this file's supplier records "
                        "may already exist in the system. Please choose another file or add new suppliers "
                        "to this file before uploading again."
                    ),
                },
            )

        return SupplierImportResult(
            created_count=len(created_suppliers),
            failed_count=len(errors),
            errors=errors,
            created_suppliers=created_suppliers,
        )

    def _validate_file(self, file: UploadFile) -> None:
        if not file.filename or not file.filename.lower().endswith(".xlsx"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only .xlsx Excel files are supported.",
            )

    def _load_excel_sheet(self, contents: bytes):
        try:
            workbook = load_workbook(BytesIO(contents), data_only=True)
        except Exception:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid Excel file.",
            )

        return workbook.active

    def _normalize_header(self, value: str) -> str:
        return value.strip().lower().replace("-", "_").replace(" ", "_")

    def _get_header_indexes(self, sheet) -> dict[str, int]:
        raw_headers = [
            self._normalize_header(str(cell.value)) if cell.value else ""
            for cell in sheet[1]
        ]

        header_indexes: dict[str, int] = {}

        for index, raw_header in enumerate(raw_headers):
            for canonical_header, aliases in HEADER_ALIASES.items():
                normalized_aliases = {
                    self._normalize_header(alias) for alias in aliases
                }

                if raw_header in normalized_aliases:
                    header_indexes[canonical_header] = index
                    break

        if "name" not in header_indexes:
            detected_headers = [header for header in raw_headers if header]

            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={
                    "message": (
                        "Could not identify the supplier name column in the uploaded Excel file."
                    ),
                    "detected_headers": detected_headers,
                    "supported_name_headers": sorted(list(HEADER_ALIASES["name"])),
                    "instructions": (
                        "Rename the supplier name column to one of the supported values "
                        "and upload the file again."
                    ),
                },
            )

        return header_indexes

    def _extract_row_data(
        self,
        row,
        header_indexes: dict[str, int],
    ) -> dict[str, str | None]:
        raw_data: dict[str, str | None] = {}

        for header, index in header_indexes.items():
            value = row[index] if index < len(row) else None
            raw_data[header] = str(value).strip() if value is not None else None

        return raw_data

    def _validate_row(
        self,
        row_number: int,
        raw_data: dict[str, str | None],
        company_id: UUID,
        seen_names: set[str],
        seen_emails: set[str],
        seen_phones: set[str],
    ) -> dict | None:
        name = raw_data.get("name")
        email = raw_data.get("email")
        phone = raw_data.get("phone")

        if not name:
             return {
                "row": row_number,
                "message": "Supplier name is missing. Please provide a supplier name in this row.",
            }

        normalized_name = name.strip()
        normalized_email = email.strip().lower() if email else None
        normalized_phone = phone.strip() if phone else None

        name_key = normalized_name.lower()

        if name_key in seen_names:
            return {
                "row": row_number,
                "message": f"Supplier '{normalized_name}' appears more than once in the uploaded Excel file.",
            }

        if normalized_email and normalized_email in seen_emails:
            return {
                "row": row_number,
                "message": f"Email '{normalized_email}' appears more than once in the uploaded Excel file.",
            }

        if normalized_phone and normalized_phone in seen_phones:
            return {
                "row": row_number,
                "message": "Duplicate supplier phone in uploaded file.",
            }

        if self.repo.get_by_name(normalized_name, company_id):
            return {
                "row": row_number,
                "message": f"Supplier '{normalized_name}' already exists in the system.",
            }

        if normalized_email and self.repo.get_by_email(normalized_email, company_id):
            return {
                "row": row_number,
                "message": f"Email '{normalized_email}' is already used by another supplier.",
            }

        if normalized_phone and self.repo.get_by_phone(normalized_phone, company_id):
           return {
                "row": row_number,
                "message": f"Phone number '{normalized_phone}' is already used by another supplier.",
            }

        return None

    def _build_supplier(
        self,
        raw_data: dict[str, str | None],
        company_id: UUID,
    ) -> Supplier:
        name = raw_data["name"].strip()
        email = raw_data.get("email")
        phone = raw_data.get("phone")
        address = raw_data.get("address")
        contact_person = raw_data.get("contact_person")

        return Supplier(
            id=uuid.uuid4(),
            company_id=company_id,
            name=name,
            email=email.strip().lower() if email else None,
            phone=phone.strip() if phone else None,
            address=address.strip() if address else None,
            contact_person=contact_person.strip() if contact_person else None,
            is_active=True,
        )